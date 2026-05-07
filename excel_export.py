"""
Генератор форматированных Excel-отчётов в стиле Россети Кубань.
"""
import io
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------- цвета (Россети) ----------
NAVY = '1F3868'
NAVY_DEEP = '142647'
BLUE = '2E5BA8'
RED = 'C8102E'
ORANGE = 'E37222'
GREEN = '0E8A6E'
LIGHT_BG = 'F4F6FA'
ZEBRA = 'FAFBFC'
RISK_BG = 'FCEBEB'
WARN_BG = 'FFF4E6'
OK_BG = 'EAF7F1'

# ---------- стили ----------
THIN = Side(border_style='thin', color='E1E5EB')
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor=NAVY)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

TITLE_FONT = Font(name='Calibri', size=18, bold=True, color=NAVY)
SUBTITLE_FONT = Font(name='Calibri', size=11, italic=True, color='5A6478')
LABEL_FONT = Font(name='Calibri', size=10, bold=True, color=NAVY)
VALUE_FONT = Font(name='Consolas', size=11, color='1A1F2C')
META_FONT = Font(name='Calibri', size=9, color='98A0AE')


def _autofit_columns(ws, max_width=50):
    """Автоширина для всех столбцов."""
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            if cell.column_letter:
                col_letter = cell.column_letter
            try:
                v = str(cell.value) if cell.value is not None else ''
                max_len = max(max_len, len(v))
            except Exception:
                pass
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


def _apply_table_format(ws, start_row, end_row, start_col, end_col,
                        zebra=True, freeze_header=True):
    """Шапка navy + зебра + границы."""
    # шапка
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=start_row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER_ALL
    ws.row_dimensions[start_row].height = 32

    # тело
    for r in range(start_row + 1, end_row + 1):
        fill = PatternFill('solid', fgColor=ZEBRA) if zebra and (r % 2 == 0) else None
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_ALL
            cell.alignment = Alignment(horizontal='left', vertical='center')
            if fill:
                cell.fill = fill

    if freeze_header:
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


# =====================================================================
# ОСНОВНАЯ ФУНКЦИЯ
# =====================================================================

def build_full_report(df_raw, registry, base, flags, report, config,
                      flag_titles, flag_emoji, metric_labels,
                      month_info):
    """Собрать полный многолистовой отчёт."""
    wb = Workbook()
    wb.remove(wb.active)

    # 1. Сводка
    _build_summary_sheet(wb, registry, flags, report, flag_titles)

    # 2. Реестр (полный)
    _build_registry_sheet(wb, registry, flags, flag_titles, flag_emoji, metric_labels)

    # 3. Потребление по абонентам с флагами
    _build_consumption_sheet(wb, df_raw, registry, flags, month_info, flag_titles, flag_emoji)

    # 4. По каждому флагу — отдельный лист
    for flag_name in flags.columns:
        if flags[flag_name].sum() > 0:
            _build_flag_sheet(wb, registry, flags, df_raw, month_info,
                              flag_name, flag_titles)

    # 5. Настройки анализа
    _build_settings_sheet(wb, config, report)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------- лист 1: Сводка ----------
def _build_summary_sheet(wb, registry, flags, report, flag_titles):
    ws = wb.create_sheet("Сводка", 0)
    ws.sheet_view.showGridLines = False

    # заголовок
    ws['B2'] = 'АНАЛИЗ АНОМАЛЬНОГО ПОТРЕБЛЕНИЯ'
    ws['B2'].font = TITLE_FONT
    ws['B3'] = f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws['B3'].font = SUBTITLE_FONT

    # навигация
    ws['B5'] = 'Период данных:'
    ws['C5'] = report.get('period_range', '—')
    ws['B6'] = 'Абонентов в файле:'
    ws['C6'] = report['total_rows']
    ws['B7'] = 'Месяцев данных:'
    ws['C7'] = report['months_detected']

    for r in range(5, 8):
        ws.cell(row=r, column=2).font = LABEL_FONT
        ws.cell(row=r, column=3).font = VALUE_FONT

    # ----- блок: общие цифры -----
    ws['B9'] = 'ОБЩИЕ ПОКАЗАТЕЛИ'
    ws['B9'].font = Font(name='Calibri', size=12, bold=True, color=NAVY)
    ws['B9'].fill = PatternFill('solid', fgColor=LIGHT_BG)
    ws.merge_cells('B9:E9')

    n_with_flags = int((registry['кол-во_флагов'] > 0).sum())
    n_high_risk = int((registry['кол-во_флагов'] >= 3).sum())
    n_clean = report['total_rows'] - n_with_flags

    summary_rows = [
        ('Без флагов', n_clean, OK_BG, GREEN),
        ('С флагами (1+)', n_with_flags, WARN_BG, ORANGE),
        ('Высокий риск (3+ флага)', n_high_risk, RISK_BG, RED),
    ]
    for i, (label, value, bg, fg) in enumerate(summary_rows):
        r = 11 + i
        ws.cell(row=r, column=2, value=label).font = Font(name='Calibri', size=11, bold=True, color=fg)
        ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor=bg)
        ws.cell(row=r, column=3, value=value).font = Font(name='Consolas', size=14, bold=True, color=fg)
        ws.cell(row=r, column=3).fill = PatternFill('solid', fgColor=bg)
        ws.cell(row=r, column=4, value=f"{value/report['total_rows']*100:.1f}%").font = Font(name='Consolas', size=11, color=fg)
        ws.cell(row=r, column=4).fill = PatternFill('solid', fgColor=bg)
        for c in range(2, 5):
            ws.cell(row=r, column=c).border = BORDER_ALL
            ws.cell(row=r, column=c).alignment = Alignment(vertical='center')

    # ----- блок: по флагам -----
    ws['B16'] = 'СРАБАТЫВАНИЕ ПО ФЛАГАМ'
    ws['B16'].font = Font(name='Calibri', size=12, bold=True, color=NAVY)
    ws['B16'].fill = PatternFill('solid', fgColor=LIGHT_BG)
    ws.merge_cells('B16:E16')

    headers = ['Флаг', 'Кол-во', 'Доля', '']
    for i, h in enumerate(headers):
        c = ws.cell(row=18, column=2 + i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        c.border = BORDER_ALL
    ws.row_dimensions[18].height = 28

    flag_data = sorted(
        [(flag_titles.get(n, n), int(flags[n].sum())) for n in flags.columns],
        key=lambda x: -x[1]
    )
    for i, (name, count) in enumerate(flag_data):
        r = 19 + i
        ws.cell(row=r, column=2, value=name).font = Font(name='Calibri', size=10)
        ws.cell(row=r, column=3, value=count).font = VALUE_FONT
        ws.cell(row=r, column=4, value=f"{count/report['total_rows']*100:.1f}%").font = VALUE_FONT
        for c in range(2, 5):
            ws.cell(row=r, column=c).border = BORDER_ALL
            if i % 2 == 0:
                ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=ZEBRA)

    # ----- блок: сезонность -----
    if 'сезонность' in registry.columns:
        seas = registry['сезонность'].value_counts()
        seas_row = 19 + len(flag_data) + 2
        ws.cell(row=seas_row, column=2, value='СЕЗОННАЯ КЛАССИФИКАЦИЯ').font = Font(name='Calibri', size=12, bold=True, color=NAVY)
        ws.cell(row=seas_row, column=2).fill = PatternFill('solid', fgColor=LIGHT_BG)
        ws.merge_cells(start_row=seas_row, start_column=2, end_row=seas_row, end_column=5)

        for i, h in enumerate(['Тип', 'Кол-во', 'Доля', '']):
            c = ws.cell(row=seas_row + 2, column=2 + i, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = HEADER_ALIGN
            c.border = BORDER_ALL

        for i, (name, count) in enumerate(seas.items()):
            r = seas_row + 3 + i
            ws.cell(row=r, column=2, value=str(name)).font = Font(name='Calibri', size=10)
            ws.cell(row=r, column=3, value=count).font = VALUE_FONT
            ws.cell(row=r, column=4, value=f"{count/report['total_rows']*100:.1f}%").font = VALUE_FONT
            for c in range(2, 5):
                ws.cell(row=r, column=c).border = BORDER_ALL
                if i % 2 == 0:
                    ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=ZEBRA)

    # ширины
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12


# ---------- лист 2: Реестр ----------
def _build_registry_sheet(wb, registry, flags, flag_titles, flag_emoji, metric_labels):
    ws = wb.create_sheet("Реестр")

    # переименуем колонки
    rename = {**metric_labels,
              **{k: f"{flag_emoji.get(k,'')} {flag_titles.get(k,k)}" for k in flags.columns}}
    df_out = registry.rename(columns=rename).copy()

    # write rows
    rows = list(dataframe_to_rows(df_out.reset_index(drop=True), index=False, header=True))
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    end_row = len(rows)
    end_col = len(df_out.columns)
    _apply_table_format(ws, start_row=1, end_row=end_row, start_col=1, end_col=end_col)

    # Подсветить строки с высоким риском (по столбцу "Кол-во флагов")
    flag_col_name = metric_labels.get('кол-во_флагов', 'кол-во_флагов')
    if flag_col_name in df_out.columns:
        flag_col_idx = list(df_out.columns).index(flag_col_name) + 1
        flag_col_letter = get_column_letter(flag_col_idx)
        # для каждой строки данных
        for r in range(2, end_row + 1):
            v = ws.cell(row=r, column=flag_col_idx).value
            if isinstance(v, (int, float)):
                if v >= 3:
                    fill = PatternFill('solid', fgColor=RISK_BG)
                elif v >= 1:
                    fill = PatternFill('solid', fgColor=WARN_BG)
                else:
                    fill = None
                if fill:
                    for c in range(1, end_col + 1):
                        ws.cell(row=r, column=c).fill = fill

    # подсветить ячейки флагов (TRUE)
    flag_display = [f"{flag_emoji.get(k,'')} {flag_titles.get(k,k)}" for k in flags.columns]
    for fname in flag_display:
        if fname in df_out.columns:
            col_idx = list(df_out.columns).index(fname) + 1
            for r in range(2, end_row + 1):
                cell = ws.cell(row=r, column=col_idx)
                if cell.value is True or cell.value == 'True':
                    cell.value = '✓'
                    cell.font = Font(name='Calibri', size=12, bold=True, color=RED)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif cell.value is False or cell.value == 'False':
                    cell.value = ''

    _autofit_columns(ws)


# ---------- лист 3: Помесячное потребление по абонентам с флагами ----------
def _build_consumption_sheet(wb, df_raw, registry, flags, month_info, flag_titles, flag_emoji):
    ws = wb.create_sheet("Потребление")

    flagged = registry[registry['кол-во_флагов'] > 0].sort_values('кол-во_флагов', ascending=False)
    if flagged.empty:
        ws['A1'] = 'Нет абонентов с флагами'
        ws['A1'].font = SUBTITLE_FONT
        return

    # заголовки
    headers = ['#', 'Кол-во флагов']
    meta_cols_check = ['Вид потребителя', 'Заводской номер прибора учета',
                       'Наименование точки учета', 'Населенный пункт',
                       'Разрешенная мощность']
    meta_cols = [c for c in meta_cols_check if c in registry.columns]
    headers.extend(meta_cols)
    month_cols = [c for (c, _, _) in month_info]
    headers.extend(month_cols)

    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _apply_table_format(ws, 1, 1 + len(flagged), 1, len(headers))

    # данные
    for i, idx in enumerate(flagged.index, start=2):
        ws.cell(row=i, column=1, value=int(idx))
        n_flags = int(flagged.loc[idx, 'кол-во_флагов'])
        ws.cell(row=i, column=2, value=n_flags)
        for j, col in enumerate(meta_cols, start=3):
            v = registry.loc[idx, col]
            ws.cell(row=i, column=j, value=v if pd.notna(v) else '')

        # потребление
        for j, mcol in enumerate(month_cols, start=3 + len(meta_cols)):
            v = pd.to_numeric(df_raw.loc[idx, mcol], errors='coerce')
            if pd.notna(v):
                cell = ws.cell(row=i, column=j, value=float(v))
                cell.number_format = '#,##0'
                cell.font = Font(name='Consolas', size=10)
                cell.alignment = Alignment(horizontal='right')

        # подсветка по риску
        if n_flags >= 3:
            fill = PatternFill('solid', fgColor=RISK_BG)
        elif n_flags >= 1:
            fill = PatternFill('solid', fgColor=WARN_BG)
        else:
            fill = None
        if fill:
            for c in range(1, 3 + len(meta_cols)):
                ws.cell(row=i, column=c).fill = fill

    # ширины: метаданные пошире, месяцы поуже
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 13
    for j, _ in enumerate(meta_cols, start=3):
        ws.column_dimensions[get_column_letter(j)].width = 22
    for j in range(3 + len(meta_cols), len(headers) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 11

    ws.freeze_panes = ws.cell(row=2, column=3 + len(meta_cols))


# ---------- лист по конкретному флагу ----------
def _build_flag_sheet(wb, registry, flags, df_raw, month_info, flag_name, flag_titles):
    """Лист с топ-кандидатами по конкретному флагу."""
    sheet_title = flag_titles.get(flag_name, flag_name)[:31]
    ws = wb.create_sheet(sheet_title)

    mask = flags[flag_name].reindex(registry.index, fill_value=False)
    sub = registry[mask].copy()
    if sub.empty:
        return

    # сортировка по приоритету
    sort_keys = {
        'подозрение_хищение': 'drop_ratio',
        'долгое_молчание': 'longest_silence',
        'превышение_мощности': 'power_ratio',
        'нестабильность': 'cv',
        'одинаковые_подряд': 'longest_flat_run',
        'аномалия_в_когорте': 'cohort_z',
    }
    sk = sort_keys.get(flag_name)
    if sk and sk in sub.columns:
        sub = sub.sort_values(sk, ascending=False)

    cols_show = ['Вид потребителя', 'Заводской номер прибора учета',
                 'Наименование точки учета', 'Населенный пункт', 'Улица', 'Дом',
                 'Разрешенная мощность', 'active_months', 'mean_kwh',
                 'avg_recent', 'avg_baseline', 'drop_ratio',
                 'longest_silence', 'longest_flat_run', 'power_ratio',
                 'cohort_z', 'сезонность', 'кол-во_флагов']
    cols_show = [c for c in cols_show if c in sub.columns]

    # заголовок листа
    ws['A1'] = f'ФЛАГ: {sheet_title.upper()}'
    ws['A1'].font = TITLE_FONT
    ws['A2'] = f'Всего абонентов: {len(sub)}'
    ws['A2'].font = SUBTITLE_FONT

    # таблица с шапкой на 4 строке
    for c, h in enumerate(cols_show, start=1):
        ws.cell(row=4, column=c, value=h)

    for i, idx in enumerate(sub.index, start=5):
        for j, col in enumerate(cols_show, start=1):
            v = sub.loc[idx, col]
            if pd.notna(v):
                if isinstance(v, (int, float, np.floating)):
                    cell = ws.cell(row=i, column=j, value=float(v))
                    if isinstance(v, float):
                        cell.number_format = '#,##0.00'
                    else:
                        cell.number_format = '#,##0'
                    cell.font = Font(name='Consolas', size=10)
                else:
                    ws.cell(row=i, column=j, value=str(v))

    _apply_table_format(ws, 4, 4 + len(sub), 1, len(cols_show))
    _autofit_columns(ws)


# ---------- лист настроек ----------
def _build_settings_sheet(wb, config, report):
    ws = wb.create_sheet("Настройки")
    ws.sheet_view.showGridLines = False

    ws['B2'] = 'ПАРАМЕТРЫ АНАЛИЗА'
    ws['B2'].font = TITLE_FONT
    ws['B3'] = 'С этими порогами получен текущий отчёт'
    ws['B3'].font = SUBTITLE_FONT

    config_labels = {
        'theft_recent_months': 'Хищение: недавний период (мес.)',
        'theft_baseline_months': 'Хищение: опорный период (мес.)',
        'theft_drop_pct': 'Хищение: порог падения (доля)',
        'theft_min_baseline_kwh': 'Хищение: мин. опорное потребление, кВт·ч',
        'silence_months': 'Молчание: подряд месяцев без показаний',
        'power_overrun_ratio': 'Мощность: порог пиковой/разрешённой',
        'cv_threshold': 'Нестабильность: порог CV',
        'flat_run_months': 'Одинаковые подряд: длина серии',
        'cohort_zscore': 'Когорта: порог robust z-score',
        'min_active_months': 'Минимум активных месяцев для анализа',
    }

    # шапка
    for c, h in enumerate(['Параметр', 'Значение'], start=2):
        cell = ws.cell(row=5, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER_ALL

    for i, (k, v) in enumerate(config.items()):
        r = 6 + i
        label = config_labels.get(k, k)
        ws.cell(row=r, column=2, value=label).font = Font(name='Calibri', size=10)
        cell_v = ws.cell(row=r, column=3, value=v)
        cell_v.font = Font(name='Consolas', size=10)
        cell_v.alignment = Alignment(horizontal='right')
        if i % 2 == 0:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=ZEBRA)
        for c in range(2, 4):
            ws.cell(row=r, column=c).border = BORDER_ALL

    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 18

    # информация о вычисленных/пропущенных флагах
    block_start = 6 + len(config) + 3
    ws.cell(row=block_start, column=2, value='Состав отчёта').font = Font(name='Calibri', size=12, bold=True, color=NAVY)
    ws.cell(row=block_start + 1, column=2,
            value=f"Вычислено флагов: {len(report.get('flags_computed', []))}").font = Font(name='Calibri', size=10)
    ws.cell(row=block_start + 2, column=2,
            value=f"Пропущено: {len(report.get('flags_skipped', []))}").font = Font(name='Calibri', size=10)
    if report.get('flags_skipped'):
        ws.cell(row=block_start + 3, column=2, value='Причины пропуска:').font = Font(name='Calibri', size=10, italic=True)
        for i, f in enumerate(report['flags_skipped']):
            ws.cell(row=block_start + 4 + i, column=2,
                    value=f"  • {f['name']}: {f['reason']}").font = Font(name='Calibri', size=9, color='5A6478')


# =====================================================================
# КАРТОЧКА АБОНЕНТА (одностраничная)
# =====================================================================

def build_subscriber_card(df_raw, registry, base, flags, idx,
                          flag_titles, flag_emoji, metric_labels, month_info):
    wb = Workbook()
    ws = wb.active
    ws.title = f'Абонент {idx}'
    ws.sheet_view.showGridLines = False

    row = registry.loc[idx]
    ws['B2'] = f'КАРТОЧКА АБОНЕНТА #{idx}'
    ws['B2'].font = TITLE_FONT

    # ----- метаданные -----
    ws['B4'] = 'МЕТАДАННЫЕ'
    ws['B4'].font = Font(name='Calibri', size=12, bold=True, color=NAVY)
    ws['B4'].fill = PatternFill('solid', fgColor=LIGHT_BG)
    ws.merge_cells('B4:D4')

    meta_keys = ['Вид потребителя', 'Заводской номер прибора учета',
                 'Наименование точки учета', 'Населенный пункт',
                 'Улица', 'Дом', 'Разрешенная мощность',
                 'Тип события', 'Дата события']
    r = 5
    for k in meta_keys:
        if k in row.index and pd.notna(row[k]):
            ws.cell(row=r, column=2, value=k).font = LABEL_FONT
            ws.cell(row=r, column=3, value=str(row[k])).font = VALUE_FONT
            r += 1

    # ----- активные флаги -----
    r += 1
    ws.cell(row=r, column=2, value='АКТИВНЫЕ ФЛАГИ').font = Font(name='Calibri', size=12, bold=True, color=NAVY)
    ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor=LIGHT_BG)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1
    active_flags = [n for n in flags.columns if flags.loc[idx, n]]
    if active_flags:
        for n in active_flags:
            ws.cell(row=r, column=2,
                    value=f"  {flag_emoji.get(n,'⚑')}  {flag_titles.get(n, n)}").font = Font(name='Calibri', size=11, color=RED, bold=True)
            ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor=RISK_BG)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
            r += 1
    else:
        ws.cell(row=r, column=2, value='  ✓  Без флагов').font = Font(name='Calibri', size=11, color=GREEN, bold=True)
        ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor=OK_BG)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 1

    # ----- ключевые метрики -----
    r += 1
    ws.cell(row=r, column=2, value='КЛЮЧЕВЫЕ МЕТРИКИ').font = Font(name='Calibri', size=12, bold=True, color=NAVY)
    ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor=LIGHT_BG)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1

    metric_keys = ['active_months', 'mean_kwh', 'avg_recent', 'avg_baseline',
                   'drop_ratio', 'longest_silence', 'longest_flat_run',
                   'power_ratio', 'cohort_z', 'сезонность', 'кол-во_флагов']
    for k in metric_keys:
        if k in row.index and pd.notna(row[k]):
            v = row[k]
            label = metric_labels.get(k, k)
            ws.cell(row=r, column=2, value=label).font = LABEL_FONT
            cell_v = ws.cell(row=r, column=3, value=v if not isinstance(v, np.floating) else float(v))
            if isinstance(v, (int, float, np.floating)) and not isinstance(v, bool):
                cell_v.number_format = '#,##0.00' if isinstance(v, float) else '#,##0'
            cell_v.font = VALUE_FONT
            r += 1

    # ----- помесячное потребление -----
    r += 1
    ws.cell(row=r, column=2, value='ПОМЕСЯЧНОЕ ПОТРЕБЛЕНИЕ, кВт·ч').font = Font(name='Calibri', size=12, bold=True, color=NAVY)
    ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor=LIGHT_BG)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1

    # шапка
    for c, h in enumerate(['Период', 'кВт·ч'], start=2):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER_ALL
    r += 1
    start_consumption = r
    for col, y, m in month_info:
        v = pd.to_numeric(df_raw.loc[idx, col], errors='coerce')
        period_label = col
        ws.cell(row=r, column=2, value=period_label).font = Font(name='Calibri', size=10)
        if pd.notna(v):
            cell_v = ws.cell(row=r, column=3, value=float(v))
            cell_v.number_format = '#,##0'
            cell_v.font = Font(name='Consolas', size=10)
            cell_v.alignment = Alignment(horizontal='right')
        if (r - start_consumption) % 2 == 0:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=ZEBRA)
        for c in range(2, 4):
            ws.cell(row=r, column=c).border = BORDER_ALL
        r += 1

    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
